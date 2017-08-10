#-*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill

'''
    ADD 
    - ACT TO EACH CASES
    - ADDITIONAL SINGULAR OR MULTIPLE (ACT, SLOT, VALUE) 
'''
# additional information excel file
book_additional = load_workbook('additional.xlsx') 
sheet_additional = book_additional.active
    
# list of all additional informations
# make values of additional_dic empty
additional_dic = {k: [] for k in range(1, 82)}  
for k in range(1, 82) : 
    additional_dic[k].extend(['', []])

# function for reading from the excel file ! 
def getInfo() :
    # get each input one by one
    for row in sheet_additional.iter_rows(min_row=2, min_col=4, max_row=100, max_col=8):
        cNum = row[0].value # case number read
        # 0 : caseNum, 1 : ACT, 2 : ACT, 3 : SLOT, 4 : VALUE
        act_1, act_2, slot, value = row[1].value, row[2].value, row[3].value, row[4].value
        additional_dic[cNum][0] = act_1
        additional_dic[cNum][1].append([act_2, slot, value])
        

''' 
    TOKENIZING PART
    INPUT AS A STRING  
'''
def insertTag (f, input, caseNum) : # dealing corpus by corpus
    if input == None :
        return
    
    print(f)
    print(input)

    list_of_tag = []
    back = -1
    tag_Num = input.count('%') # the number of tag in a corpus
    natural = ''

    
    for i in range(tag_Num) :
        # get necessary indexes
        tag_index_num = input.index('%', back+1)
        front = input.index('{', back+1)
        natural += input[back+1:front] 
        back = input.index('}', back+1)
        
        # get act, slot and value
        act = additional_dic[caseNum][0]
        value = input[front+1:tag_index_num]
        slot = input[tag_index_num+1:back]
        
        # insert act, slot, and value into the list
        list_of_tag.append([act, slot, value])
        
        # making natural one of the input
        natural += value 
        if ( i == tag_Num - 1 ) : 
            natural += input[back+1:]
    
    if (tag_Num == 0) :
        natural += input
        
    corp_info_list = []
    if ( len(additional_dic[caseNum][1]) > 0 ) :
        list_of_tag.extend(additional_dic[caseNum][1])
        
    corp_info_list.extend([input, natural, list_of_tag, caseNum])
    all_corpus.append(corp_info_list)

    
'''
    MAIN STARTS HERE
'''
# make list for the corpus information
files = ['1.xlsx', '2.xlsx', '3.xlsx', '4.xlsx', '5.xlsx', '6.xlsx', '7.xlsx', '8.xlsx'
         , '9.xlsx', '10.xlsx', '11.xlsx', '12.xlsx', '13.xlsx', '14.xlsx', '15.xlsx', '16.xlsx', '17.xlsx'
         , '18.xlsx']


# write book 
write_book = Workbook()
write_sheet = write_book.active

# additional information 
getInfo()
index = 1
# get each input one by one
for f in files : 
    all_corpus = []
# read from the corpus excel file
    book = load_workbook(f)
    sheet = book.active
    caseNum = 0
    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=33):
        caseNum += 1
        for cell in row:
            input = cell.value 
            # process
            insertTag(f, input, caseNum)
        
    # write to a new excel file
    for i in all_corpus :
        leng = len(i[2])
        first_flag = 1
        for j in i[2] :
            if ( first_flag ) :  
                row = (str(index), i[0], i[1], j[0], j[1], j[2])
                first_flag = 0
                index += 1 
            else :
                row = ('', '', '', j[0], j[1], j[2])
        
            write_sheet.append(row)
        
    # write_sheet.merge_cells('A' + str(cur_row-leng) + ':A' + str(cur_row-1))
    # write_sheet.merge_cells('B' + str(cur_row-leng) + ':B' + str(cur_row-1))


'''
    EXCEL CELL STYLE PART
'''
# adjust cell size
dims = {}
for row in write_sheet.rows:
    for cell in row:
        if cell.value:
            dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
for col, value in dims.items():
    write_sheet.column_dimensions[col].width = value
    
# cell alignment
for row in write_sheet.rows : 
    for cell in row :
        cell.alignment = Alignment(horizontal="justify")

write_book.save('tagged_corpus.xlsx')


